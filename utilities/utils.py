import softest
import logging
import inspect
import csv
from openpyxl import load_workbook


class Utils(softest.TestCase):

    def assert_text(self, list_items, value):
        print("\nTotal Flights with desired stops: ", len(list_items))
        for stop in list_items:
            print("The text is: " + stop.text)
            self.soft_assert(self.assertEqual, stop.text, value)
            if stop.text == value:
                print("test passed")
            else:
                print("test failed")
        self.assert_all()

    def add_logs(log_level=logging.DEBUG):
        # Set class/method name from where it is called
        logger_name = inspect.stack()[1][3]
        # Create logger
        logger = logging.getLogger(logger_name)
        # Set log level
        logger.setLevel(log_level)
        # Create file handler
        file_handler = logging.FileHandler("automation.log", mode="a")
        # Create Formatter
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')
        # Add formatter to file handler
        file_handler.setFormatter(formatter)
        # add handler to logger
        logger.addHandler(file_handler)
        return logger

    def read_data_from_excel(file_name, sheet):
        data = []
        wb = load_workbook(filename=file_name)
        sheet = wb[sheet]
        num_rows = sheet.max_row
        num_cols = sheet.max_column

        for i in range(2, num_rows + 1):
            row = []
            for j in range(1, num_cols + 1):
                row.append(sheet.cell(row=i, column=j).value)
            data.append(row)
        return data

    def read_data_from_csv(file_name):
        data = []
        with open(file_name, 'r') as file:
            reader = csv.reader(file)
            next(reader)  # header
            for row in reader:
                data.append(row)
        return data
